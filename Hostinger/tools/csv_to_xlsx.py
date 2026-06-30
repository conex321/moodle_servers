"""Convert domains/hostinger_domains.csv into a domains-only Excel workbook.

Usage: python csv_to_xlsx.py
Reads ../domains/hostinger_domains.csv and writes ../domains/hostinger_domains.xlsx
with a single "Domains" sheet: bold frozen header row, auto-sized columns, autofilter.
"""
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "..", "domains", "hostinger_domains.csv")
DST = os.path.join(HERE, "..", "domains", "hostinger_domains.xlsx")

with open(SRC, newline="", encoding="utf-8") as f:
    rows = list(csv.reader(f))

if not rows:
    raise SystemExit("CSV is empty: " + SRC)

header, data = rows[0], rows[1:]

wb = Workbook()
ws = wb.active
ws.title = "Domains"

ws.append(header)
for r in data:
    ws.append(r)

# Style header
bold = Font(bold=True)
for cell in ws[1]:
    cell.font = bold
    cell.alignment = Alignment(vertical="center")

# Freeze header + autofilter
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

# Auto-size columns (cap width for readability)
for col_idx, col_name in enumerate(header, start=1):
    max_len = len(str(col_name))
    for r in data:
        if col_idx - 1 < len(r):
            max_len = max(max_len, len(str(r[col_idx - 1])))
    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)

wb.save(DST)
print(f"WROTE {DST}: {len(data)} domains, {len(header)} columns")
